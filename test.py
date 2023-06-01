import os
from openpyxl import Workbook, load_workbook
import openpyxl
import openai

openai.api_type = "azure"
openai.api_base = "https://openaipoc-edt.openai.azure.com/"
openai.api_version = "2023-03-15-preview"
openai.api_key = "2eab1ea1f1d741f4ba1c71c45c9313d2"

softwareCategory = ""

def store_info():

        # Define the headers
        headers = ['Software/Application Name', 'Category',
                'Top 5 Features', 'Short History', 'Version History']

        filename = "final.xlsx"

        if os.path.isfile(filename):
            # If the Excel file already exists, load it
            workbook = load_workbook(filename=filename)
            sheet = workbook.active
        else:
            # If the Excel file doesn't exist, create it and add the headers
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(headers)

        # Append the software information to the sheet
        software_info = [softwareName, softwareCategory]
        sheet.append(software_info)

        # Save the workbook
        workbook.save(filename=filename)


def generateResponse():
        response = openai.ChatCompletion.create(
        engine="EDTOpenAIgptturbo",
        messages = [{"role":"system","content":"You are an AI assistant that helps people find information about software. Provide a 1 word answer."},{"role":"user","content":"What category of software is " + softwareName}],
        temperature=0.7,
        max_tokens=800,
        top_p=0.95,
        frequency_penalty=0,
        presence_penalty=0,
        stop=None)
        print(response.choices[0].message.content)
        softwareCategory = response.choices[0].message.content
     
     



# open the excel file
wb = openpyxl.load_workbook('start.xlsx')

# get the active sheet
sheet = wb.active

# loop through the rows
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    # get the value of the first column
    softwareName = row[0].value
    # save the values to variables
    softwareName = softwareName
    generateResponse()
    store_info()

print("Complete!")