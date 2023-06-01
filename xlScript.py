import os
from openpyxl import Workbook, load_workbook

# open the excel file
wb = openpyxl.load_workbook('example.xlsx')

# get the active sheet
sheet = wb.active

# loop through the rows
for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
    # get the value of the first column
    softwareName = row[0].value
    # get the value of the second column
    softwareT1 = row[1].value
    # save the values to variables
    softwareName = softwareName
    softwareT1 = softwareT1

def store_info():
    # Ask the user for the software information
    software_name = input("Enter Software/Application Name: ")
    category = input("Enter Category: ")
    features = input("Enter Top 5 Features (comma-separated): ").split(',')
    history = input("Enter Short History: ")
    version_history = input("Enter Version History: ")

    # Define the headers
    headers = ['Software/Application Name', 'Category',
               'Top 5 Features', 'Short History', 'Version History']

    filename = "Software_Info.xlsx"

    if os.path.isfile(filename):
        # If the Excel file already exists, load it
        workbook = load_workbook(filename=filename)
        sheet = workbook.active
    else:
        # If the Excel file doesn't exist, create it and add the headers
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)

    # Append the software information to the sheet
    software_info = [software_name, category,
                     ', '.join(features), history, version_history]
    sheet.append(software_info)

    # Save the workbook
    workbook.save(filename=filename)


store_info()
